# -*- coding: utf-8 -*-
# Developed by Youssef Omri AKA DZEUF

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import requests
import logging
from openpyxl import load_workbook
from io import BytesIO

_logger = logging.getLogger(__name__)


class ImportProductWizard(models.TransientModel):
    _name = 'import.product.wizard'
    _description = 'Import Product Wizard'

    file = fields.Binary('Excel File', required=True)
    filename = fields.Char('Filename')

    def action_import_products(self):
        if not self.file:
            raise UserError(_("Please upload a valid Excel file."))

        try:
            file_data = base64.b64decode(self.file)
            workbook = load_workbook(filename=BytesIO(file_data), data_only=True)
            sheet = workbook.active  # First sheet

            queue = self.env['import.product.queue'].create({})
            row_count = 0

            for row in sheet.iter_rows(min_row=3, values_only=True):  # Start from 3rd row (1-based index)
                # Skip row if all relevant fields are empty or critical ones (like SKU/Name) are missing
                if not any([row[0], row[1], row[2], row[3], row[7], row[10]]) or not row[1] or not row[2]:
                    continue

                category_text = row[0]
                sku = row[1]
                name = row[2]
                ean = row[3]
                brand = row[7]
                image_url = row[10]

                category_id = None
                if category_text:
                    last_word = category_text.split('/')[-1].strip()
                    category_id = self.env['product.category'].search([('name', '=ilike', last_word)], limit=1)

                image_data = None
                if image_url:
                    try:
                        img_response = requests.get(image_url, timeout=10)
                        if img_response.status_code == 200:
                            image_data = base64.b64encode(img_response.content)
                    except Exception as e:
                        _logger.warning("Failed to fetch image for SKU %s: %s", sku, str(e))

                self.env['import.product.queue.line'].create({
                    'queue_id': queue.id,
                    'category_id': category_id.id if category_id else False,
                    'sku': sku,
                    'name': name,
                    'ean': ean,
                    'brand': brand,
                    'main_image': image_data,
                })
                row_count += 1

            self.env['bus.bus']._sendone(self.env.user.partner_id, 'simple_notification', {
                'type': 'success',
                'message': _("Product import completed. %d lines created.") % row_count,
                'sticky': False,
            })

            self.env['import.product.queue.log'].create({
                'queue_id': queue.id,
                'message': f"Imported {row_count} product lines from Excel.",
                'status': 'success'
            })

        except Exception as e:
            _logger.exception("Error during product import")
            raise UserError(_("Import failed: %s") % str(e))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Product Queue'),
            'res_model': 'import.product.queue',
            'view_mode': 'form',
            'res_id': queue.id,
            'target': 'current',
        }
